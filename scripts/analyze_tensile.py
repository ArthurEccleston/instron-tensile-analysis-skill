"""
Instron Tensile Test Mechanical Properties Analyzer

Scans a folder for *-*.is_tens_Exports directories, reads CSV files inside,
extracts tensile strength (max 拉伸应力) and elongation at break (断裂伸长率).

Output: 力学性能_YYYYMMDD.xlsx with two sheets:
  - 原始数据: long-format raw data (one row per sample, ready for plotting)
  - 统计汇总: per-type mean, std with outlier removal
"""

import sys
import os
import re
import csv
import statistics
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ---- CSV helpers ----

def _open_csv(filepath):
    for enc in ("utf-8-sig", "gbk", "gb2312", "gb18030", "utf-8"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                csv.reader(f).__next__()
            return enc
        except (UnicodeDecodeError, StopIteration):
            continue
    return "utf-8-sig"


def get_csv_headers(filepath):
    enc = _open_csv(filepath)
    with open(filepath, "r", encoding=enc) as f:
        return next(csv.reader(f))


def find_col_index(headers, target):
    for idx, h in enumerate(headers, start=1):
        if target in h:
            return idx
    return None


def read_csv_values(filepath, col_idx, start_row=3):
    enc = _open_csv(filepath)
    values = []
    with open(filepath, "r", encoding=enc) as f:
        reader = csv.reader(f)
        for row_num, row in enumerate(reader, start=1):
            if row_num < start_row:
                continue
            if col_idx <= len(row):
                val = row[col_idx - 1].strip().strip('"')
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    continue
    return values


# ---- Mechanical property extraction ----

def operation_x(filepath):
    """Max tensile stress (抗拉强度)."""
    headers = get_csv_headers(filepath)
    col = find_col_index(headers, "拉伸应力")
    if col is None:
        return None
    values = read_csv_values(filepath, col)
    return max(values) if values else None


def operation_v(filepath):
    """
    Elongation at break (断裂伸长率):
    Detect fracture by locating the first CATASTROPHIC stress drop after the
    peak stress, then return the strain value just before that drop.

    Algorithm:
    1. Locate max stress row
    2. Compute stress-diff baseline from pre-peak region
    3. Scan forward from the peak: find the first diff that is BOTH
       statistically anomalous AND exceeds the absolute drop threshold
    4. Return the strain at the row just before that drop
    """
    WINDOW = 30       # rows before peak to establish stress-diff baseline
    K = 6.0           # sigma threshold
    MIN_DROP = -1.0   # absolute minimum stress drop (MPa) to count as fracture

    headers = get_csv_headers(filepath)
    stress_col = find_col_index(headers, "拉伸应力")
    strain_col = find_col_index(headers, "拉伸应变")
    if stress_col is None or strain_col is None:
        print(f"  WARNING: '拉伸应力' or '拉伸应变' column not found")
        return None

    stress_vals = read_csv_values(filepath, stress_col)
    strain_vals = read_csv_values(filepath, strain_col)
    if len(stress_vals) < WINDOW + 2:
        return strain_vals[-1] if strain_vals else None

    # Find peak stress position
    max_idx = max(range(len(stress_vals)), key=lambda i: stress_vals[i])

    # Baseline: stress diffs from a window before the peak
    baseline_start = max(0, max_idx - WINDOW)
    baseline_diffs = [
        stress_vals[i + 1] - stress_vals[i]
        for i in range(baseline_start, max_idx - 1)
    ]
    if len(baseline_diffs) < 3:
        return strain_vals[max_idx]

    baseline_mean = statistics.mean(baseline_diffs)
    baseline_std = statistics.stdev(baseline_diffs)
    threshold = baseline_mean - K * baseline_std
    print(f"  [V] baseline mean_diff={baseline_mean:.4f} std={baseline_std:.4f} "
          f"threshold={threshold:.4f} | min_abs_drop={MIN_DROP}")

    # Scan forward from the peak for a catastrophic drop
    for i in range(max_idx, len(stress_vals) - 1):
        diff = stress_vals[i + 1] - stress_vals[i]
        if diff < threshold and diff < MIN_DROP:
            print(f"  [V] Fracture at row {i + 3}: "
                  f"stress_diff={diff:.2f} < {threshold:.2f} AND < {MIN_DROP}")
            return strain_vals[i]

    # Fallback: no catastrophic drop found, use strain at peak stress
    print(f"  [V] No catastrophic drop found; using strain at peak stress")
    return strain_vals[max_idx]


# ---- File scanning ----

def parse_dirname(dirname):
    name = dirname.replace(".is_tens_Exports", "")
    m = re.match(r"^(\d+)[-_](\d+)$", name)
    return (int(m.group(1)), int(m.group(2))) if m else None


def scan_dirs(folder):
    folder_path = Path(folder)
    if not folder_path.is_dir():
        print(f"ERROR: '{folder}' is not a valid directory.")
        sys.exit(1)
    results = []
    for entry in sorted(folder_path.iterdir()):
        if not entry.is_dir() or ".is_tens_Exports" not in entry.name:
            continue
        csv_files = sorted(entry.glob("*.csv"))
        if not csv_files:
            print(f"SKIP: '{entry.name}' — no CSV files inside")
            continue
        results.append((entry, csv_files[0]))
    return results


# ---- Outlier removal (IQR method) ----

def remove_outliers(values):
    """
    Remove outliers using the 1.5*IQR rule.
    Only applied when there are >= 4 values; smaller sets are returned as-is.
    Returns (cleaned_values, removed_indices).
    """
    if len(values) < 4:
        return values, []

    sorted_v = sorted(values)
    n = len(sorted_v)
    q1 = sorted_v[n // 4]
    q3 = sorted_v[(3 * n) // 4]
    iqr = q3 - q1
    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr

    cleaned = []
    removed = []
    for i, v in enumerate(values):
        if lower <= v <= upper:
            cleaned.append(v)
        else:
            removed.append((i, v))
    return cleaned, removed


# ---- Style helpers ----

def style_header(ws, row, ncols):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")


def style_data_rows(ws, start_row, end_row, ncols):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")


# ---- Main ----

def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_tensile.py <folder_path> [output_dir]")
        sys.exit(1)

    folder = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else folder
    date_str = datetime.now().strftime("%Y%m%d")
    output_path = os.path.join(output_dir, f"力学性能_{date_str}.xlsx")

    entries = scan_dirs(folder)
    if not entries:
        print("No .is_tens_Exports directories found.")
        sys.exit(0)

    print(f"Found {len(entries)} directories to process.\n")

    # Collect raw data: list of (type, num, tensile_strength, elongation)
    raw_data = []

    for dir_path, csv_path in entries:
        parsed = parse_dirname(dir_path.name)
        if parsed is None:
            print(f"SKIP: '{dir_path.name}' — cannot parse type-num from name")
            continue
        sample_type, sample_num = parsed
        print(f"Processing: {dir_path.name}/ -> {csv_path.name}  (type={sample_type}, num={sample_num})")

        ts = operation_x(str(csv_path))
        el = operation_v(str(csv_path))

        if ts is not None:
            print(f"  [X] 抗拉强度: {ts:.2f} MPa")
        else:
            print(f"  [X] 抗拉强度: FAILED")
        if el is not None:
            print(f"  [V] 断裂伸长率: {el:.2f} %")
        else:
            print(f"  [V] 断裂伸长率: FAILED")

        raw_data.append({
            "type": sample_type,
            "num": sample_num,
            "tensile": ts,
            "elongation": el,
        })
        print()

    # Group by type
    types = sorted(set(d["type"] for d in raw_data))
    groups = {}
    for t in types:
        groups[t] = [d for d in raw_data if d["type"] == t]

    # ============================================================
    # Build output workbook
    # ============================================================
    wb = Workbook()

    # ---- Sheet 1: 原始数据 ----
    ws1 = wb.active
    ws1.title = "原始数据"
    headers1 = ["种类", "编号", "抗拉强度(MPa)", "断裂伸长率(%)"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    row_idx = 2
    for d in raw_data:
        ws1.cell(row=row_idx, column=1, value=d["type"])
        ws1.cell(row=row_idx, column=2, value=d["num"])
        if d["tensile"] is not None:
            ws1.cell(row=row_idx, column=3, value=round(d["tensile"], 2))
        if d["elongation"] is not None:
            ws1.cell(row=row_idx, column=4, value=round(d["elongation"], 2))
        row_idx += 1

    style_data_rows(ws1, 2, row_idx - 1, len(headers1))
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18

    # ---- Sheet 2: 统计汇总 ----
    ws2 = wb.create_sheet("统计汇总")
    headers2 = [
        "种类", "原始样品数", "有效样品数",
        "抗拉强度均值(MPa)", "抗拉强度标准差(MPa)",
        "断裂伸长率均值(%)", "断裂伸长率标准差(%)",
        "剔除异常值",
    ]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    note_font = Font(color="FF0000", size=10)

    row_idx = 2
    for t in types:
        group = groups[t]
        raw_tensile = [d["tensile"] for d in group if d["tensile"] is not None]
        raw_elong = [d["elongation"] for d in group if d["elongation"] is not None]

        clean_tensile, removed_t = remove_outliers(raw_tensile)
        clean_elong, removed_e = remove_outliers(raw_elong)

        # Gather outlier notes
        outlier_notes = []
        for idx, val in removed_t:
            # Find the sample num
            sample = group[idx]
            outlier_notes.append(f"#{sample['num']} 抗拉={val:.2f}")
        for idx, val in removed_e:
            sample = group[idx]
            outlier_notes.append(f"#{sample['num']} 伸长率={val:.2f}")

        ws2.cell(row=row_idx, column=1, value=t)
        ws2.cell(row=row_idx, column=2, value=len(raw_tensile))
        ws2.cell(row=row_idx, column=3, value=len(clean_tensile))

        if len(clean_tensile) >= 1:
            ws2.cell(row=row_idx, column=4, value=round(statistics.mean(clean_tensile), 2))
            ws2.cell(row=row_idx, column=5, value=round(statistics.stdev(clean_tensile), 2) if len(clean_tensile) >= 2 else "-")
        else:
            ws2.cell(row=row_idx, column=4, value="-")
            ws2.cell(row=row_idx, column=5, value="-")

        if len(clean_elong) >= 1:
            ws2.cell(row=row_idx, column=6, value=round(statistics.mean(clean_elong), 2))
            ws2.cell(row=row_idx, column=7, value=round(statistics.stdev(clean_elong), 2) if len(clean_elong) >= 2 else "-")
        else:
            ws2.cell(row=row_idx, column=6, value="-")
            ws2.cell(row=row_idx, column=7, value="-")

        note = "; ".join(outlier_notes) if outlier_notes else "无"
        cell_note = ws2.cell(row=row_idx, column=8, value=note)
        if outlier_notes:
            cell_note.font = note_font

        row_idx += 1

    style_data_rows(ws2, 2, row_idx - 1, len(headers2))
    for col_letter, width in [("A", 8), ("B", 14), ("C", 14), ("D", 20), ("E", 22), ("F", 20), ("G", 22), ("H", 50)]:
        ws2.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f"Done. {len(raw_data)} samples processed across {len(types)} types.")
    print(f"Output: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
